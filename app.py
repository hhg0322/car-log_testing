import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
from datetime import datetime

st.title("🚗 차량 운행일지 자동 생성기")

# 1. 사용자 입력 폼
with st.form("log_form"):
    st.subheader("운행 기록 추가")
    col1, col2 = st.columns(2)
    
    date = col1.date_input("날짜", value=datetime.now())
    start_time = col2.text_input("출발시간 (예: 13:00)", "13:00")
    end_time = col1.text_input("도착시간 (예: 14:00)", "14:00")
    start_km = col2.text_input("운행 전 Km", "17666")
    distance = col1.text_input("주행거리(Km)", "111")
    end_km = col2.text_input("운행 후 Km", "17777")
    destination = col1.text_input("경유지 및 목적지", "화성, 화서역")
    category = col2.text_input("운행목적", "오후 정기운행")
    refuel_cost = col1.text_input("주유(원)", "0")
    refuel_liter = col2.text_input("주유량(L)", "0")
    
    submitted = st.form_submit_button("기록 추가하기")
    
    if submitted:
        new_log = {
            "date": str(date),
            "startTime": start_time,
            "endTime": end_time,
            "startKm": start_km,
            "distance": distance,
            "endKm": end_km,
            "destination": destination,
            "category": category,
            "refuelCost": refuel_cost,
            "refuelLiter": refuel_liter
        }
        if 'logs' not in st.session_state:
            st.session_state.logs = []
        st.session_state.logs.append(new_log)
        st.success("기록이 임시 추가되었습니다!")

# 2. 입력된 목록 확인 및 엑셀 다운로드 기능
if 'logs' in st.session_state and st.session_state.logs:
    st.markdown("---")
    st.subheader("📋 입력된 운행 기록 목록")
    st.dataframe(pd.DataFrame(st.session_state.logs))

    if st.button("🗑️ 목록 초기화"):
        st.session_state.logs = []
        st.rerun()

    # 엑셀 파일 생성 함수
    def generate_excel(data):
        wb = load_workbook('car_log_template.xlsx', data_only=True)
        ws = wb.active
        
        # 월 자동 기입 (C1)
        ws['C1'] = f"{datetime.now().month:02d}월 차량운행일지"
        
        def safe_set(coord, val):
            if type(ws[coord]).__name__ != 'MergedCell':
                ws[coord].value = val

        for idx, log in enumerate(data):
            r = 6 + idx
            if r > 15: break
            safe_set(f'A{r}', idx + 1)
            safe_set(f'C{r}', log['date'])
            safe_set(f'D{r}', log['startTime'])
            safe_set(f'E{r}', log['endTime'])
            safe_set(f'F{r}', log['startKm'])
            safe_set(f'G{r}', log['distance'])
            safe_set(f'H{r}', log['endKm'])
            safe_set(f'I{r}', log['destination'])
            safe_set(f'J{r}', log['category'])
            safe_set(f'K{r}', log['refuelCost'])
            safe_set(f'N{r}', None) # 서명란 비워둠
            
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # 엑셀 다운로드 버튼
    excel_file = generate_excel(st.session_state.logs)
    st.download_button(
        label="📥 엑셀 파일 다운로드 받기",
        data=excel_file,
        file_name="car_log_result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
